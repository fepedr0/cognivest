from flask import Flask, request, jsonify, render_template, redirect, session, send_file
from flask_mysqldb import MySQL
from flask_bcrypt import Bcrypt
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl
import io
import os

app = Flask(__name__, static_folder="assets")
app.secret_key = os.environ.get('SECRET_KEY', 'cognivest_secret_123')

app.config['MYSQL_HOST'] = os.environ.get('MYSQL_HOST', 'localhost')
app.config['MYSQL_USER'] = os.environ.get('MYSQL_USER', 'root')
app.config['MYSQL_PASSWORD'] = os.environ.get('MYSQL_PASSWORD', 'felipe2008hg')
app.config['MYSQL_DB'] = os.environ.get('MYSQL_DB', 'cognivest')
app.config['MYSQL_PORT'] = int(os.environ.get('MYSQL_PORT', 3306))

mysql = MySQL(app)
bcrypt = Bcrypt(app)

@app.route('/')
def home():
    return redirect('/login')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('login.html')
    
    dados = request.get_json()
    email = dados.get('email')
    senha = dados.get('senha')

    cur = mysql.connection.cursor()
    cur.execute("SELECT id, nome, senha_hash FROM usuarios WHERE email = %s", (email,))
    usuario = cur.fetchone()
    cur.close()

    if usuario and bcrypt.check_password_hash(usuario[2], senha):
        session['usuario_id'] = usuario[0]
        session['usuario_nome'] = usuario[1]
        return jsonify({ 'sucesso': True })
    
    return jsonify({ 'sucesso': False, 'erro': 'Email ou senha incorretos' })

@app.route('/cadastro', methods=['GET', 'POST'])
def cadastro():
    if request.method == 'GET':
        return render_template('cadastro.html')

    dados = request.get_json()
    nome = dados.get('nome')
    email = dados.get('email')
    senha = dados.get('senha')

    senha_hash = bcrypt.generate_password_hash(senha).decode('utf-8')

    try:
        cur = mysql.connection.cursor()
        cur.execute("INSERT INTO usuarios (nome, email, senha_hash) VALUES (%s, %s, %s)", 
                    (nome, email, senha_hash))
        mysql.connection.commit()
        cur.close()
        return jsonify({ 'sucesso': True })
    except Exception as e:
        print('ERRO NO CADASTRO:', e)
        return jsonify({ 'sucesso': False, 'erro': 'Email já cadastrado' })

@app.route('/calculadora')
def calculadora():
    if 'usuario_id' not in session:
        return redirect('/login')
    return render_template('index.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

@app.route('/salvar_calculo', methods=['POST'])
def salvar_calculo():
    if 'usuario_id' not in session:
        return jsonify({ 'sucesso': False, 'erro': 'Não autenticado' })

    dados = request.get_json()

    try:
        cur = mysql.connection.cursor()
        cur.execute("""
            INSERT INTO calculos 
            (usuario_id, tipo_bem, valor_bem, taxa_juros, parcelas, entrada, parcela_mensal, valor_final, total_juros, aumento, vale_a_pena)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            session['usuario_id'],
            dados['tipo_bem'],
            dados['valor_bem'],
            dados['taxa_juros'],
            dados['parcelas'],
            dados['entrada'],
            dados['parcela_mensal'],
            dados['valor_final'],
            dados['total_juros'],
            dados['aumento'],
            dados['vale_a_pena']
        ))
        mysql.connection.commit()
        cur.close()
        return jsonify({ 'sucesso': True })
    except Exception as e:
        print('ERRO AO SALVAR:', e)
        return jsonify({ 'sucesso': False })

@app.route('/historico')
def historico():
    if 'usuario_id' not in session:
        return redirect('/login')
    
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM calculos WHERE usuario_id = %s ORDER BY criado_em DESC", (session['usuario_id'],))
    calculos = cur.fetchall()
    cur.close()

    return render_template('historico.html', calculos=calculos)

@app.route('/gerar_planilha', methods=['POST'])
def gerar_planilha():
    if 'usuario_id' not in session:
        return jsonify({ 'sucesso': False, 'erro': 'Não autenticado' })

    dados = request.get_json()
    tipo_bem = dados['tipo_bem']
    valor_bem = float(dados['valor_bem'])
    taxa_juros = float(dados['taxa_juros'])
    parcelas = int(dados['parcelas'])
    entrada = float(dados['entrada'])

    financed = valor_bem - entrada
    rate = taxa_juros / 100
    parcela = financed * (rate * (1 + rate) ** parcelas) / ((1 + rate) ** parcelas - 1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financiamento"

    ws.append(["Mês", "Parcela (R$)", "Juros (R$)", "Amortização (R$)", "Saldo Devedor (R$)"])

    header_fill = PatternFill("solid", fgColor="6d28d9")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    saldo = financed
    for mes in range(1, parcelas + 1):
        juros_mes = saldo * rate
        amortizacao = parcela - juros_mes
        saldo -= amortizacao
        ws.append([
            mes,
            round(parcela, 2),
            round(juros_mes, 2),
            round(amortizacao, 2),
            round(max(saldo, 0), 2)
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'cognivest_{tipo_bem}.xlsx'
    )

@app.route('/limpar_historico', methods=['POST'])
def limpar_historico():
    if 'usuario_id' not in session:
        return jsonify({ 'sucesso': False })
    
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM calculos WHERE usuario_id = %s", (session['usuario_id'],))
    mysql.connection.commit()
    cur.close()
    
    return jsonify({ 'sucesso': True })

if __name__ == '__main__':
    app.run(debug=True)